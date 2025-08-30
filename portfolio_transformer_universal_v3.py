"""
Enhanced Portfolio Transformer Module - Universal Multi-Format Support with Fixed Detection

Handles three input formats:
1. Original BOD Excel format (Day Beginning positions)
2. CSV format (POSITION_DETAILS_NEW with Contract Id parsing)
3. MS Position Excel format (Contract in Col 1, Position in Col 22)

Features:
- Improved format detection based on column headers
- Auto-loads "futures mapping.csv"
- Auto-fetches Yahoo Finance prices
- Password protection handling for Excel files
- Fund selection (Aurigin/Wafra)
- Bloomberg price integration
- Flexible start row detection for all formats

Author: Portfolio Management System
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Union
import logging
from dataclasses import dataclass
import yfinance as yf
import warnings
import re
import os

# Suppress yfinance warnings
warnings.filterwarnings("ignore", category=FutureWarning)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def select_fund() -> str:
    """Prompt user to select fund"""
    print("🦁 FUND SELECTION")
    print("=" * 30)
    print("1. Aurigin")
    print("2. Wafra")
    
    while True:
        try:
            choice = input("\nSelect fund (1 or 2): ").strip()
            if choice == '1':
                print("✅ Selected: Aurigin")
                return "Aurigin"
            elif choice == '2':
                print("✅ Selected: Wafra")
                return "Wafra"
            else:
                print("❌ Invalid choice. Please enter 1 or 2.")
        except KeyboardInterrupt:
            raise SystemExit("\nUser cancelled selection")


def select_file_from_directory(file_type: str, extensions: List[str]) -> str:
    """Show user available files and let them select one"""
    print(f"\n=== {file_type.upper()} FILE SELECTION ===")
    
    # Get all files with specified extensions
    available_files = []
    for file in os.listdir("."):
        if any(file.lower().endswith(ext.lower()) for ext in extensions):
            available_files.append(file)
    
    if not available_files:
        print(f"❌ No {file_type} files found with extensions: {', '.join(extensions)}")
        manual_input = input(f"Enter {file_type} filename manually (or 'quit' to exit): ").strip()
        if manual_input.lower() == 'quit':
            raise SystemExit("User chose to quit")
        return manual_input
    
    # Display available files
    print(f"Available {file_type} files:")
    for i, file in enumerate(available_files, 1):
        file_size = os.path.getsize(file) / 1024  # KB
        print(f"  {i}. {file} ({file_size:.1f} KB)")
    
    # Get user selection
    while True:
        try:
            choice = input(f"\nSelect {file_type} file (1-{len(available_files)}) or enter filename directly: ").strip()
            
            # Check if it's a number
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(available_files):
                    selected_file = available_files[idx]
                    print(f"✅ Selected: {selected_file}")
                    return selected_file
                else:
                    print(f"❌ Invalid selection. Please choose 1-{len(available_files)}")
                    continue
            
            # Check if it's a direct filename
            elif choice in available_files:
                print(f"✅ Selected: {choice}")
                return choice
            
            # Check if file exists (manual entry)
            elif os.path.exists(choice):
                print(f"✅ Selected: {choice}")
                return choice
            
            else:
                print(f"❌ File not found: {choice}")
                print("Please try again or choose from the numbered list.")
                
        except KeyboardInterrupt:
            raise SystemExit("\nUser cancelled selection")
        except Exception as e:
            print(f"❌ Error: {e}")


def read_excel_with_password(file_path: str, **kwargs) -> pd.DataFrame:
    """Try to read Excel file with password handling"""
    passwords = ["Aurigin2017", "Aurigin2024"]
    
    # First try without password
    try:
        return pd.read_excel(file_path, **kwargs)
    except Exception as e:
        logger.info(f"File appears to be password protected, trying default passwords...")
        
        # Try default passwords
        for password in passwords:
            try:
                logger.info(f"Trying password: {password}")
                import openpyxl
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path, password=password)
                ws = wb.active
                
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    nrows = kwargs.get('nrows', None)
                    header = kwargs.get('header', 0)
                    
                    if header is None:
                        # No header, return all data as is
                        if nrows and len(data) > nrows:
                            data = data[:nrows]
                        df = pd.DataFrame(data)
                    else:
                        # With header
                        if nrows and len(data) > nrows + 1:
                            data = data[:nrows + 1]
                        if len(data) > 0:
                            df = pd.DataFrame(data[1:], columns=data[0])
                        else:
                            df = pd.DataFrame()
                    
                    logger.info(f"✅ Successfully opened with password: {password}")
                    return df
                    
            except Exception as pwd_error:
                logger.debug(f"Password {password} failed: {str(pwd_error)}")
                continue
        
        # If all default passwords fail, ask user
        while True:
            try:
                user_password = input(f"\n🔒 Default passwords failed. Enter password for {file_path} (or 'skip' to skip): ").strip()
                if user_password.lower() == 'skip':
                    raise Exception("User chose to skip password-protected file")
                
                wb = load_workbook(file_path, password=user_password)
                ws = wb.active
                
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    nrows = kwargs.get('nrows', None)
                    header = kwargs.get('header', 0)
                    
                    if header is None:
                        if nrows and len(data) > nrows:
                            data = data[:nrows]
                        df = pd.DataFrame(data)
                    else:
                        if nrows and len(data) > nrows + 1:
                            data = data[:nrows + 1]
                        if len(data) > 0:
                            df = pd.DataFrame(data[1:], columns=data[0])
                        else:
                            df = pd.DataFrame()
                    
                    logger.info(f"✅ Successfully opened with user password")
                    return df
                    
            except KeyboardInterrupt:
                raise SystemExit("\nUser cancelled")
            except Exception as e:
                print(f"❌ Password failed: {str(e)}")
                continue


@dataclass
class Position:
    """Data class for position information"""
    underlying_ticker: str
    symbol: str
    bloomberg_ticker: str
    series: str
    expiry: datetime
    strike: float
    option_type: str
    position: float
    lot_size: int
    security_type: str
    deliverable: float
    underlying_price: Optional[float] = None
    override_price: Optional[float] = None
    bbg_price: Optional[float] = None


class EnhancedPortfolioTransformer:
    """Enhanced class for transforming multiple input formats to deliverable format"""
    
    def __init__(self, fund_name: str):
        self.fund_name = fund_name
        self.mapping_data = {}
        self.positions = []
        self.unmapped_positions = []  # Track positions without mappings
        self.underlying_prices = {}
        self.price_overrides = {}
        self.bbg_price_overrides = {}
        self.input_format = None
    
    def load_mapping_data(self, mapping_file_path: str = "futures mapping.csv") -> None:
        """Load symbol mapping from CSV file"""
        try:
            if not os.path.exists(mapping_file_path):
                raise FileNotFoundError(f"Mapping file not found: {mapping_file_path}")
            
            df = pd.read_csv(mapping_file_path)
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            # Use the first 3 columns as Symbol, Ticker, Cash
            if len(df.columns) >= 3:
                df = df.iloc[:, :3]
                df.columns = ['Symbol', 'Ticker', 'Cash']
            
            # Build mapping dictionary - strip spaces from symbols
            for _, row in df.iterrows():
                if pd.notna(row['Symbol']) and pd.notna(row['Ticker']):
                    symbol = str(row['Symbol']).strip()
                    ticker = str(row['Ticker']).strip()
                    cash = str(row['Cash']).strip() if pd.notna(row['Cash']) else f"{ticker} IS Equity"
                    
                    self.mapping_data[symbol] = {
                        'ticker': ticker,
                        'cash_ticker': cash
                    }
            
            logger.info(f"✅ Loaded {len(self.mapping_data)} symbol mappings from {mapping_file_path}")
            
        except Exception as e:
            logger.error(f"Error loading mapping file: {str(e)}")
            raise
    
    def _detect_excel_format(self, df_raw: pd.DataFrame, file_path: str) -> str:
        """Enhanced format detection for Excel files with flexible row detection"""
        
        logger.info("🔍 Detecting file format...")
        
        # Check 1: MS Position format - specific pattern check
        # MS Position has contracts in column 1 AND position data in column 22
        if len(df_raw.columns) >= 22:
            ms_position_likely = False
            for row_idx in range(5, min(25, len(df_raw))):
                col1_val = str(df_raw.iloc[row_idx, 0]).strip() if pd.notna(df_raw.iloc[row_idx, 0]) else ""
                # Check if column 1 has contract patterns (FUTSTK-SYMBOL-DATE-TYPE-STRIKE)
                if '-' in col1_val and any(x in col1_val.upper() for x in ['FUTSTK', 'OPTSTK']):
                    # Also check if column 22 has numeric data
                    if len(df_raw.columns) > 21:
                        try:
                            col22_val = float(df_raw.iloc[row_idx, 21])
                            if not pd.isna(col22_val):
                                ms_position_likely = True
                                break
                        except:
                            pass
            
            if ms_position_likely:
                logger.info("🔍 MS Position format detected: contracts in col 1, numeric data in col 22")
                return 'excel_ms_position'
        
        # Check 2: BOD format detection - look for symbols in column 2 and position in column 16
        # This should be checked BEFORE looking for "Open Position" text
        if len(df_raw.columns) >= 16:
            bod_likely = False
            for row_idx in range(5, min(30, len(df_raw))):
                if len(df_raw.iloc[row_idx]) < 16:
                    continue
                
                # Check column 2 for symbol-like values
                col2_val = str(df_raw.iloc[row_idx, 1]).strip() if pd.notna(df_raw.iloc[row_idx, 1]) else ""
                
                # BOD format indicators:
                # - Column 2 has alphabetic symbols (stock tickers)
                # - Column 3 might have FUTSTK/OPTSTK or be empty
                # - Column 16 has numeric position data
                if col2_val and len(col2_val) >= 2 and col2_val.replace('&', '').replace('-', '').isalpha():
                    # Check if this could be a symbol
                    if col2_val in self.mapping_data or col2_val.upper() == col2_val:
                        # Also check column 16 for numeric data
                        try:
                            if pd.notna(df_raw.iloc[row_idx, 15]):
                                pos_val = float(df_raw.iloc[row_idx, 15])
                                bod_likely = True
                                logger.info(f"🔍 BOD format detected: symbol '{col2_val}' in col 2, position data in col 16")
                                break
                        except:
                            pass
            
            if bod_likely:
                return 'excel_bod'
        
        # Check 3: Look for Contract Id column header (Format 2)
        try:
            df_with_headers = read_excel_with_password(file_path, nrows=5)
            column_names = [str(col).lower() for col in df_with_headers.columns]
            if any('contract' in col and 'id' in col for col in column_names):
                logger.info("🔍 Found 'Contract Id' column header - Format 2 detected")
                return 'excel_contract_id'
        except:
            pass
        
        # Check 4: File name hints - check these BEFORE defaulting
        file_lower = file_path.lower()
        if 'ms' in file_lower and 'position' in file_lower:
            logger.info("🔍 Filename suggests MS Position format")
            return 'excel_ms_position'
        elif 'bod' in file_lower or 'beginning' in file_lower or 'day' in file_lower:
            logger.info("🔍 Filename suggests BOD format")
            return 'excel_bod'
        elif 'contract' in file_lower:
            logger.info("🔍 Filename suggests Contract ID format")
            return 'excel_contract_id'
        
        # Default to BOD format
        logger.info("🔍 No specific format indicators found, defaulting to BOD format")
        return 'excel_bod'
    
    def load_positions(self, input_file_path: str, start_row: int = 12) -> None:
        """Auto-detect input format and load positions accordingly"""
        try:
            file_ext = os.path.splitext(input_file_path.lower())[1]
            
            if file_ext == '.csv':
                self.input_format = 'csv_contract_id'
                self._load_csv_positions(input_file_path)
            elif file_ext in ['.xlsx', '.xls']:
                # Read file to analyze structure
                try:
                    # Read without headers first to see raw data
                    df_raw = read_excel_with_password(input_file_path, header=None, nrows=30)
                    
                    logger.info(f"📊 Analyzing Excel file: {df_raw.shape[0]} rows, {df_raw.shape[1]} columns")
                    
                    # Use enhanced detection logic
                    detected_format = self._detect_excel_format(df_raw, input_file_path)
                    
                    if detected_format == 'excel_contract_id':
                        self.input_format = 'excel_contract_id'
                        logger.info("✅ Detected Format 2: Excel with Contract Id column")
                        self._load_excel_contract_positions(input_file_path)
                    elif detected_format == 'excel_ms_position':
                        self.input_format = 'excel_ms_position'
                        logger.info("✅ Detected Format 3: MS Position sheet")
                        self._load_ms_position_positions(input_file_path)
                    else:
                        self.input_format = 'excel_bod'
                        logger.info("✅ Detected Format 1: Original BOD Excel")
                        self._load_bod_positions(input_file_path, start_row)
                        
                except Exception as e:
                    logger.error(f"Could not read Excel file: {str(e)}")
                    raise
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            logger.info(f"✅ Final detected format: {self.input_format}")
            logger.info(f"✅ Loaded {len(self.positions)} positions")
            
        except Exception as e:
            logger.error(f"Error loading positions file: {str(e)}")
            raise
    
    def _load_csv_positions(self, csv_file_path: str) -> None:
        """Load positions from CSV format with flexible column detection"""
        try:
            df = pd.read_csv(csv_file_path)
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            logger.info(f"📊 CSV format - Analyzing {df.shape[0]} rows, {df.shape[1]} columns")
            logger.info(f"📋 Column headers: {df.columns.tolist()[:10]}...")
            
            # Find columns flexibly
            contract_id_col = None
            cf_lots_col = None
            lot_size_col = None
            
            # Look for contract ID column
            for col in df.columns:
                col_lower = col.lower()
                if 'contract' in col_lower and ('id' in col_lower or 'code' in col_lower):
                    contract_id_col = col
                    break
            
            # Look for position/lots column
            for col in df.columns:
                col_lower = col.lower()
                if ('cf' in col_lower and 'lot' in col_lower) or ('carry' in col_lower and 'forward' in col_lower):
                    cf_lots_col = col
                    break
            
            # Look for lot size column
            for col in df.columns:
                col_lower = col.lower()
                if 'lot' in col_lower and 'size' in col_lower:
                    lot_size_col = col
                    break
            
            # If not found by name, try by position and pattern
            if not contract_id_col:
                # Look for column with contract-like patterns
                for col in df.columns:
                    sample_values = df[col].dropna().astype(str).head(5)
                    if any('-' in val and len(val) > 20 for val in sample_values):
                        contract_id_col = col
                        logger.info(f"✅ Found contract column by pattern: {col}")
                        break
                
                # Fallback to position
                if not contract_id_col and len(df.columns) >= 4:
                    contract_id_col = df.columns[3]
                    logger.info(f"⚠️ Using column position 4 for contracts: {contract_id_col}")
            
            if not cf_lots_col and len(df.columns) >= 11:
                cf_lots_col = df.columns[10]
                logger.info(f"⚠️ Using column position 11 for lots: {cf_lots_col}")
            
            if not lot_size_col and len(df.columns) >= 6:
                lot_size_col = df.columns[5]
                logger.info(f"⚠️ Using column position 6 for lot size: {lot_size_col}")
            
            if not all([contract_id_col, cf_lots_col, lot_size_col]):
                missing = []
                if not contract_id_col: missing.append("Contract ID")
                if not cf_lots_col: missing.append("CF Lots")
                if not lot_size_col: missing.append("Lot Size")
                raise ValueError(f"Could not identify required columns: {missing}")
            
            positions = []
            unmapped_positions = []
            skipped_zero = 0
            skipped_parse_error = 0
            skipped_no_mapping = 0
            processed = 0
            
            for idx, row in df.iterrows():
                try:
                    contract_id = str(row[contract_id_col]).strip()
                    cf_lots = float(row[cf_lots_col]) if pd.notna(row[cf_lots_col]) else 0.0
                    lot_size = int(row[lot_size_col]) if pd.notna(row[lot_size_col]) else 1
                    
                    # Skip zero positions
                    if cf_lots == 0:
                        skipped_zero += 1
                        continue
                    
                    # Parse contract ID
                    parsed_data = self._parse_contract_id(contract_id)
                    if not parsed_data:
                        logger.warning(f"Could not parse contract ID: {contract_id}")
                        skipped_parse_error += 1
                        continue
                    
                    symbol = parsed_data['symbol']
                    
                    # Check mapping
                    if symbol not in self.mapping_data:
                        logger.warning(f"No mapping found for symbol: '{symbol}' (row {idx+2})")
                        skipped_no_mapping += 1
                        # Store unmapped position details
                        unmapped_positions.append({
                            'symbol': symbol,
                            'contract_id': contract_id,
                            'position': cf_lots,
                            'lot_size': lot_size,
                            'series': parsed_data['series'],
                            'expiry': parsed_data['expiry'],
                            'strike': parsed_data['strike'],
                            'option_type': parsed_data['option_type'],
                            'row_number': idx + 2
                        })
                        continue
                    
                    mapping_info = self.mapping_data[symbol]
                    cash_ticker = mapping_info['cash_ticker']
                    fo_ticker = mapping_info['ticker']
                    
                    # Generate Bloomberg ticker
                    bloomberg_ticker = self._generate_bloomberg_ticker(
                        symbol, fo_ticker, parsed_data['series'], parsed_data['expiry'], 
                        parsed_data['strike'], parsed_data['option_type']
                    )
                    
                    # Determine security type
                    security_type = self._get_security_type(parsed_data['series'], parsed_data['option_type'])
                    
                    position = Position(
                        underlying_ticker=cash_ticker,
                        symbol=symbol,
                        bloomberg_ticker=bloomberg_ticker,
                        series=parsed_data['series'],
                        expiry=parsed_data['expiry'],
                        strike=parsed_data['strike'],
                        option_type=parsed_data['option_type'],
                        position=cf_lots,
                        lot_size=lot_size,
                        security_type=security_type,
                        deliverable=0.0
                    )
                    
                    positions.append(position)
                    processed += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing row {idx+2}: {str(e)}")
                    continue
            
            self.positions = positions
            self.unmapped_positions = unmapped_positions
            self._print_processing_summary(processed, skipped_zero, skipped_parse_error, skipped_no_mapping, len(df))
            
        except Exception as e:
            logger.error(f"Error loading CSV file: {str(e)}")
            raise
    
    def _load_excel_contract_positions(self, excel_file_path: str) -> None:
        """Load positions from Excel file with Contract Id format - flexible detection"""
        try:
            df = read_excel_with_password(excel_file_path)
            
            # Clean column names
            df.columns = df.columns.str.strip()
            
            logger.info(f"📊 Excel Contract format - Analyzing {df.shape[0]} rows, {df.shape[1]} columns")
            logger.info(f"📋 Column headers: {df.columns.tolist()[:10]}...")
            
            # Find required columns flexibly
            contract_col = None
            lots_col = None
            lot_size_col = None
            
            # Search for contract column
            for col in df.columns:
                col_lower = col.lower()
                if 'contract' in col_lower and ('id' in col_lower or 'code' in col_lower):
                    contract_col = col
                    logger.info(f"✅ Found contract column: {col}")
                    break
            
            # Search for lots column
            for col in df.columns:
                col_lower = col.lower()
                if ('cf' in col_lower and 'lot' in col_lower) or ('carry' in col_lower and 'forward' in col_lower):
                    lots_col = col
                    logger.info(f"✅ Found lots column: {col}")
                    break
            
            # Search for lot size column
            for col in df.columns:
                col_lower = col.lower()
                if 'lot' in col_lower and 'size' in col_lower:
                    lot_size_col = col
                    logger.info(f"✅ Found lot size column: {col}")
                    break
            
            # If not found by name, look for patterns
            if not contract_col:
                for col in df.columns:
                    sample_values = df[col].dropna().astype(str).head(10)
                    # Check if column contains contract-like patterns
                    contract_count = sum(1 for val in sample_values if '-' in val and len(val) > 20)
                    if contract_count >= 3:  # At least 3 contract-like values
                        contract_col = col
                        logger.info(f"✅ Found contract column by pattern: {col}")
                        break
            
            # Fallback to positions if columns not found
            if not all([contract_col, lots_col, lot_size_col]):
                if len(df.columns) >= 11:
                    if not contract_col:
                        contract_col = df.columns[3]
                        logger.info(f"⚠️ Using position 4 for contract: {contract_col}")
                    if not lots_col:
                        lots_col = df.columns[10]
                        logger.info(f"⚠️ Using position 11 for lots: {lots_col}")
                    if not lot_size_col:
                        lot_size_col = df.columns[5]
                        logger.info(f"⚠️ Using position 6 for lot size: {lot_size_col}")
                else:
                    missing = []
                    if not contract_col: missing.append("Contract")
                    if not lots_col: missing.append("Lots")
                    if not lot_size_col: missing.append("Lot Size")
                    raise ValueError(f"Could not identify required columns: {missing}")
            
            positions = []
            skipped_zero = 0
            skipped_parse_error = 0
            skipped_no_mapping = 0
            processed = 0
            
            for idx, row in df.iterrows():
                try:
                    contract_id = str(row[contract_col]).strip()
                    cf_lots = float(row[lots_col]) if pd.notna(row[lots_col]) else 0.0
                    lot_size = int(row[lot_size_col]) if pd.notna(row[lot_size_col]) else 1
                    
                    if cf_lots == 0:
                        skipped_zero += 1
                        continue
                    
                    parsed_data = self._parse_contract_id(contract_id)
                    if not parsed_data:
                        skipped_parse_error += 1
                        continue
                    
                    symbol = parsed_data['symbol']
                    
                    if symbol not in self.mapping_data:
                        logger.warning(f"No mapping found for symbol: '{symbol}' (row {idx+2})")
                        skipped_no_mapping += 1
                        # Store unmapped position details
                        self.unmapped_positions.append({
                            'symbol': symbol,
                            'contract_id': contract_id,
                            'position': cf_lots,
                            'lot_size': lot_size,
                            'series': parsed_data['series'],
                            'expiry': parsed_data['expiry'],
                            'strike': parsed_data['strike'],
                            'option_type': parsed_data['option_type'],
                            'row_number': idx + 2,
                            'source': 'Excel Contract Format'
                        })
                        continue
                    
                    mapping_info = self.mapping_data[symbol]
                    cash_ticker = mapping_info['cash_ticker']
                    fo_ticker = mapping_info['ticker']
                    
                    bloomberg_ticker = self._generate_bloomberg_ticker(
                        symbol, fo_ticker, parsed_data['series'], 
                        parsed_data['expiry'], parsed_data['strike'], parsed_data['option_type']
                    )
                    
                    security_type = self._get_security_type(parsed_data['series'], parsed_data['option_type'])
                    
                    position = Position(
                        underlying_ticker=cash_ticker,
                        symbol=symbol,
                        bloomberg_ticker=bloomberg_ticker,
                        series=parsed_data['series'],
                        expiry=parsed_data['expiry'],
                        strike=parsed_data['strike'],
                        option_type=parsed_data['option_type'],
                        position=cf_lots,
                        lot_size=lot_size,
                        security_type=security_type,
                        deliverable=0.0
                    )
                    
                    positions.append(position)
                    processed += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing Excel row {idx+2}: {str(e)}")
                    continue
            
            self.positions = positions
            self._print_processing_summary(processed, skipped_zero, skipped_parse_error, skipped_no_mapping, len(df))
            
        except Exception as e:
            logger.error(f"Error loading Excel contract file: {str(e)}")
            raise
    
    def _load_ms_position_positions(self, excel_file_path: str) -> None:
        """Load positions from MS Position sheet Excel format"""
        try:
            # First, read raw data to find where actual data starts
            df_raw = read_excel_with_password(excel_file_path, header=None)
            
            logger.info(f"📊 MS Position format - Analyzing {df_raw.shape[0]} rows, {df_raw.shape[1]} columns")
            
            # Find the header row containing "Open Position" or similar
            header_row_idx = None
            position_col_idx = None
            
            for row_idx in range(min(20, len(df_raw))):
                row_values = df_raw.iloc[row_idx].fillna('').astype(str)
                for col_idx, val in enumerate(row_values):
                    val_lower = str(val).lower()
                    if 'open' in val_lower and ('position' in val_lower or 'pos' in val_lower):
                        header_row_idx = row_idx
                        position_col_idx = col_idx
                        logger.info(f"✅ Found header row at index {header_row_idx} with 'Open Position' at column {position_col_idx + 1}")
                        break
                if header_row_idx is not None:
                    break
            
            # If no header found, assume standard structure
            if header_row_idx is None:
                logger.info("⚠️ No explicit header found, assuming standard MS Position structure")
                header_row_idx = 0
                position_col_idx = 21  # Column 22 (0-indexed)
                data_start_row = 1
            else:
                data_start_row = header_row_idx + 1
            
            # Now process the data starting from the correct row
            positions = []
            skipped_zero = 0
            skipped_parse_error = 0
            skipped_no_mapping = 0
            processed = 0
            
            logger.info(f"📋 Processing data starting from row {data_start_row + 1}")
            
            for idx in range(data_start_row, len(df_raw)):
                try:
                    row = df_raw.iloc[idx]
                    
                    # Contract in Column 1 (index 0)
                    contract_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    
                    # Position in the identified column (or column 22)
                    if position_col_idx is not None and len(row) > position_col_idx:
                        position_value = row.iloc[position_col_idx]
                    elif len(row) > 21:  # Fallback to column 22
                        position_value = row.iloc[21]
                    else:
                        position_value = 0
                    
                    # Try to convert position to float
                    try:
                        open_pos = float(position_value) if pd.notna(position_value) else 0.0
                    except (ValueError, TypeError):
                        # Skip non-numeric values (like additional headers)
                        continue
                    
                    # Skip if no valid contract or zero position
                    if not contract_id or contract_id.lower() in ['nan', 'none', ''] or '-' not in contract_id:
                        if open_pos != 0:
                            logger.debug(f"Skipping row {idx + 1}: invalid contract '{contract_id}'")
                        continue
                    
                    if open_pos == 0:
                        skipped_zero += 1
                        continue
                    
                    # Parse contract ID
                    parsed_data = self._parse_contract_id(contract_id)
                    if not parsed_data:
                        logger.warning(f"Could not parse contract ID: '{contract_id}' (row {idx + 1})")
                        skipped_parse_error += 1
                        continue
                    
                    symbol = parsed_data['symbol']
                    
                    # Check mapping
                    if symbol not in self.mapping_data:
                        logger.warning(f"No mapping found for symbol: '{symbol}' (row {idx + 1})")
                        skipped_no_mapping += 1
                        # Store unmapped position details
                        self.unmapped_positions.append({
                            'symbol': symbol,
                            'contract_id': contract_id,
                            'position': open_pos,
                            'lot_size': 1,
                            'series': parsed_data['series'],
                            'expiry': parsed_data['expiry'],
                            'strike': parsed_data['strike'],
                            'option_type': parsed_data['option_type'],
                            'row_number': idx + 1,
                            'source': 'MS Position Format'
                        })
                        continue
                    
                    mapping_info = self.mapping_data[symbol]
                    cash_ticker = mapping_info['cash_ticker']
                    fo_ticker = mapping_info['ticker']
                    
                    # Generate Bloomberg ticker
                    bloomberg_ticker = self._generate_bloomberg_ticker(
                        symbol, fo_ticker, parsed_data['series'], 
                        parsed_data['expiry'], parsed_data['strike'], parsed_data['option_type']
                    )
                    
                    # Determine security type
                    security_type = self._get_security_type(parsed_data['series'], parsed_data['option_type'])
                    
                    # Create position
                    position = Position(
                        underlying_ticker=cash_ticker,
                        symbol=symbol,
                        bloomberg_ticker=bloomberg_ticker,
                        series=parsed_data['series'],
                        expiry=parsed_data['expiry'],
                        strike=parsed_data['strike'],
                        option_type=parsed_data['option_type'],
                        position=open_pos,
                        lot_size=1,  # Default lot size for MS Position format
                        security_type=security_type,
                        deliverable=0.0
                    )
                    
                    positions.append(position)
                    processed += 1
                    
                    logger.debug(f"✅ Processed: {contract_id} -> {symbol} with position {open_pos}")
                    
                except Exception as e:
                    logger.warning(f"Error processing MS Position row {idx + 1}: {str(e)}")
                    continue
            
            self.positions = positions
            self._print_processing_summary(processed, skipped_zero, skipped_parse_error, skipped_no_mapping, len(df_raw) - data_start_row)
            
        except Exception as e:
            logger.error(f"Error loading MS Position file: {str(e)}")
            raise
    
    def _load_bod_positions(self, bod_file_path: str, start_row: int = 12) -> None:
        """Load BOD positions from Excel file with flexible start row detection"""
        try:
            df = read_excel_with_password(bod_file_path, header=None)
            
            logger.info(f"📊 BOD format - Analyzing {df.shape[0]} rows, {df.shape[1]} columns")
            
            # Find where data actually starts by looking for valid symbols in column 2
            data_start_row = None
            for i in range(min(30, len(df))):  # Check first 30 rows
                if len(df.iloc[i]) < 16:  # Need at least 16 columns
                    continue
                    
                # Check if column 2 has a valid symbol (should be in mapping)
                col2_val = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
                
                # Check if it looks like a symbol and exists in mapping
                if col2_val and col2_val in self.mapping_data:
                    # Also verify column 16 has numeric data (position)
                    try:
                        pos_val = float(df.iloc[i, 15]) if pd.notna(df.iloc[i, 15]) else 0
                        if isinstance(pos_val, (int, float)):
                            data_start_row = i
                            logger.info(f"✅ Found data starting at row {data_start_row + 1}")
                            break
                    except:
                        continue
            
            # If no valid data found with mapping, try pattern matching
            if data_start_row is None:
                for i in range(min(30, len(df))):
                    if len(df.iloc[i]) < 16:
                        continue
                    
                    col2_val = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
                    col3_val = str(df.iloc[i, 2]).strip() if pd.notna(df.iloc[i, 2]) else ""
                    
                    # Look for patterns suggesting this is data row
                    if (col2_val and len(col2_val) > 1 and col2_val.isalpha() and 
                        col3_val in ['FUTSTK', 'OPTSTK', 'EQ', '']):
                        data_start_row = i
                        logger.info(f"✅ Found data starting at row {data_start_row + 1} (pattern match)")
                        break
            
            # Fallback to default
            if data_start_row is None:
                data_start_row = start_row
                logger.info(f"⚠️ Using default start row: {data_start_row + 1}")
            
            # Extract data starting from detected row
            data_df = df.iloc[data_start_row:].copy()
            
            positions = []
            skipped_zero = 0
            skipped_no_mapping = 0
            skipped_invalid = 0
            processed = 0
            
            for idx, row in data_df.iterrows():
                # Skip empty rows or rows with insufficient data
                if len(row) < 16 or pd.isna(row.iloc[1]):
                    continue
                
                try:
                    symbol = str(row.iloc[1]).strip()
                    
                    # Skip if symbol looks like a header
                    if symbol.lower() in ['symbol', 'ticker', 'stock', 'underlying']:
                        continue
                    
                    series = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                    expiry_raw = row.iloc[3]
                    strike = float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0.0
                    option_type = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
                    lot_size = int(row.iloc[6]) if pd.notna(row.iloc[6]) else 1
                    open_position = float(row.iloc[15]) if pd.notna(row.iloc[15]) else 0.0
                    
                    # Parse expiry date
                    expiry = self._parse_date(expiry_raw)
                    
                    # Skip positions with zero lots
                    if open_position == 0:
                        skipped_zero += 1
                        continue
                    
                    # Get mapping info
                    if symbol not in self.mapping_data:
                        logger.warning(f"No mapping found for symbol: {symbol} (row {idx + 1})")
                        skipped_no_mapping += 1
                        # Store unmapped position details
                        self.unmapped_positions.append({
                            'symbol': symbol,
                            'contract_id': f"{symbol}-{series}-{expiry.strftime('%d%b%Y').upper()}-{option_type}-{strike}",
                            'position': open_position,
                            'lot_size': lot_size,
                            'series': series,
                            'expiry': expiry,
                            'strike': strike,
                            'option_type': option_type,
                            'row_number': idx + 1,
                            'source': 'BOD Format'
                        })
                        continue
                    
                    mapping_info = self.mapping_data[symbol]
                    cash_ticker = mapping_info['cash_ticker']
                    fo_ticker = mapping_info['ticker']
                    
                    # Generate Bloomberg ticker
                    bloomberg_ticker = self._generate_bloomberg_ticker(
                        symbol, fo_ticker, series, expiry, strike, option_type
                    )
                    
                    # Determine security type
                    security_type = self._get_security_type(series, option_type)
                    
                    position = Position(
                        underlying_ticker=cash_ticker,
                        symbol=symbol,
                        bloomberg_ticker=bloomberg_ticker,
                        series=series,
                        expiry=expiry,
                        strike=strike,
                        option_type=option_type,
                        position=open_position,
                        lot_size=lot_size,
                        security_type=security_type,
                        deliverable=0.0
                    )
                    
                    positions.append(position)
                    processed += 1
                    
                except Exception as e:
                    logger.warning(f"Error processing BOD row {idx + 1}: {str(e)}")
                    skipped_invalid += 1
                    continue
            
            self.positions = positions
            self._print_processing_summary(processed, skipped_zero, skipped_invalid, skipped_no_mapping, len(data_df))
            
        except Exception as e:
            logger.error(f"Error loading BOD file: {str(e)}")
            raise
    
    def _print_processing_summary(self, processed, skipped_zero, skipped_parse_error, skipped_no_mapping, total_rows):
        """Print processing summary"""
        logger.info(f"📊 Processing Summary:")
        logger.info(f"  ✅ Processed successfully: {processed}")
        logger.info(f"  ⭕ Skipped (zero positions): {skipped_zero}")
        logger.info(f"  ❌ Skipped (parse errors): {skipped_parse_error}")
        logger.info(f"  🔍 Skipped (no mapping): {skipped_no_mapping}")
        logger.info(f"  📋 Total rows analyzed: {total_rows}")
        
        # Add unmapped positions count
        if hasattr(self, 'unmapped_positions') and self.unmapped_positions:
            logger.info(f"  ⚠️ Unmapped positions stored: {len(self.unmapped_positions)}")
            unique_unmapped = set(pos['symbol'] for pos in self.unmapped_positions)
            logger.info(f"  📝 Unique unmapped symbols: {len(unique_unmapped)}")
            logger.info(f"     First few: {list(unique_unmapped)[:5]}")
    
    def _parse_contract_id(self, contract_id: str) -> Optional[Dict]:
        """Parse contract ID string to extract components"""
        try:
            parts = contract_id.split('-')
            
            if len(parts) < 5:
                return None
            
            contract_type = parts[0].strip()
            symbol = parts[1].strip()
            date_str = parts[2].strip()
            option_type = parts[3].strip()
            strike_str = parts[4].strip()
            
            # Parse expiry date
            expiry = self._parse_date_string(date_str)
            
            # Parse strike
            strike = float(strike_str) if strike_str else 0.0
            
            # Determine series
            if contract_type == 'FUTSTK':
                series = 'FUTSTK'
            elif contract_type == 'OPTSTK':
                series = 'OPTSTK'
            else:
                series = 'UNKNOWN'
            
            return {
                'symbol': symbol,
                'expiry': expiry,
                'option_type': option_type,
                'strike': strike,
                'series': series,
                'contract_type': contract_type
            }
            
        except Exception as e:
            return None
    
    def _parse_date_string(self, date_str: str) -> datetime:
        """Parse date string like '28AUG2025' to datetime"""
        try:
            # Common month abbreviations
            month_map = {
                'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
            }
            
            # Extract day, month, year using regex
            match = re.match(r'(\d{1,2})([A-Z]{3})(\d{4})', date_str.upper())
            if match:
                day = int(match.group(1))
                month_abbr = match.group(2)
                year = int(match.group(3))
                
                month = month_map.get(month_abbr)
                if month:
                    return datetime(year, month, day)
            
            # Fallback: try pandas parsing
            return pd.to_datetime(date_str)
            
        except Exception as e:
            logger.warning(f"Could not parse date {date_str}, using today: {str(e)}")
            return datetime.now()
    
    def fetch_underlying_prices(self, symbols: List[str] = None) -> Dict[str, float]:
        """Fetch underlying prices from Yahoo Finance with enhanced retry logic"""
        if symbols is None:
            symbols = list(set(pos.symbol for pos in self.positions))
        
        logger.info(f"📈 Fetching Yahoo Finance prices for {len(symbols)} symbols...")
        prices = {}
        
        for symbol in symbols:
            try:
                # Try different Yahoo Finance symbol formats for Indian stocks
                yahoo_symbols = [
                    f"{symbol}.NS",  # NSE format
                    f"{symbol}.BO",  # BSE format
                    symbol           # Direct symbol
                ]
                
                price = None
                for yahoo_symbol in yahoo_symbols:
                    try:
                        ticker = yf.Ticker(yahoo_symbol)
                        hist = ticker.history(period="1d")
                        
                        if not hist.empty:
                            price = hist['Close'].iloc[-1]
                            logger.info(f"✅ Fetched price for {symbol}: {price} (via {yahoo_symbol})")
                            break
                    except:
                        continue
                
                if price is not None:
                    prices[symbol] = float(price)
                else:
                    logger.warning(f"❌ Could not fetch price for {symbol}")
                    
            except Exception as e:
                logger.warning(f"Error fetching price for {symbol}: {str(e)}")
        
        self.underlying_prices.update(prices)
        logger.info(f"✅ Successfully fetched {len(prices)} prices")
        return prices
    
    def calculate_deliverables(self, auto_fetch_prices: bool = True) -> None:
        """Calculate deliverable positions based on moneyness"""
        
        # Auto-fetch prices
        if auto_fetch_prices and not self.underlying_prices:
            self.fetch_underlying_prices()
        
        logger.info("🧮 Calculating deliverables...")
        
        for position in self.positions:
            try:
                # Get system-fetched price
                system_price = self.underlying_prices.get(position.symbol)
                
                # Store prices for output
                position.underlying_price = system_price
                
                # Calculate deliverable based on position type
                if position.security_type == "Futures":
                    # Futures: deliverable = position (1:1)
                    position.deliverable = position.position
                    
                elif position.security_type in ["Call", "Put"]:
                    # Options: depends on moneyness
                    if system_price is None:
                        logger.warning(f"No price available for {position.underlying_ticker}, assuming ITM")
                        is_itm = True  # Default assumption
                    else:
                        is_itm = self._is_in_the_money(
                            position.option_type, position.strike, system_price
                        )
                    
                    if is_itm:
                        if position.security_type == "Call":
                            position.deliverable = position.position
                        else:  # Put
                            position.deliverable = -position.position
                    else:
                        position.deliverable = 0.0
                else:
                    position.deliverable = 0.0
                    logger.warning(f"Unknown security type for {position.symbol}: {position.security_type}")
                
            except Exception as e:
                logger.error(f"Error calculating deliverable for {position.symbol}: {str(e)}")
                position.deliverable = 0.0
    
    def save_output_excel(self, output_path: str) -> None:
        """Save output to Excel file with formulas, grouping, and unmapped symbols sheet"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill
            
            # Debug logging
            logger.info(f"📊 Saving output: {len(self.positions)} mapped positions, {len(self.unmapped_positions)} unmapped positions")
            
            if not self.positions and not self.unmapped_positions:
                logger.warning("No positions to export")
                return
            
            # Create workbook and remove default sheet
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Create sheets for mapped positions if any exist
            if self.positions:
                # Create Net Position Summary sheet FIRST
                self._create_net_position_summary(wb, self.positions)
                logger.info(f"✅ Created Net Position Summary sheet")
                
                # Create Price Alert sheet
                self._create_price_alert_sheet(wb, self.positions)
                logger.info(f"✅ Created Price Alert sheet")
                
                # Group positions by expiry date
                expiry_groups = {}
                for position in self.positions:
                    expiry_key = position.expiry.strftime('%Y-%m-%d')
                    if expiry_key not in expiry_groups:
                        expiry_groups[expiry_key] = []
                    expiry_groups[expiry_key].append(position)
                
                # Create master sheet with ALL positions
                self._create_grouped_sheet(wb, "Master_All_Expiries", self.positions)
                logger.info(f"✅ Created Master sheet with {len(self.positions)} positions")
                
                # Create individual expiry sheets
                for expiry_date, positions in sorted(expiry_groups.items()):
                    sheet_name = f"Expiry_{expiry_date.replace('-', '_')}"
                    if len(sheet_name) > 31:
                        sheet_name = f"Exp_{expiry_date.replace('-', '_')}"
                    
                    self._create_grouped_sheet(wb, sheet_name, positions)
                    logger.info(f"✅ Created sheet '{sheet_name}' with {len(positions)} positions")
            
            # Create unmapped symbols sheet if there are any
            if self.unmapped_positions:
                logger.info(f"⚠️ Creating unmapped sheet with {len(self.unmapped_positions)} positions")
                self._create_unmapped_sheet(wb)
                logger.info(f"✅ Created 'Unmapped_Symbols' sheet with {len(self.unmapped_positions)} positions")
            else:
                logger.info("ℹ️ No unmapped positions found - all symbols had mappings")
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"💾 Excel output saved to: {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            raise
    
    def _create_net_position_summary(self, workbook, positions: List) -> None:
        """Create net position summary sheet showing deliverables by underlying"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title="Net_Position_Summary", index=0)
        
        # Headers
        headers = [
            'Underlying', 'Total Contracts', 'Total Lots', 'Lot Size',
            'System Deliverable', 'Override Deliverable', 'BBG Deliverable',
            'System Price', 'Override Price', 'BBG Price'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Aggregate positions by underlying
        underlying_summary = {}
        for position in positions:
            underlying = position.underlying_ticker
            if underlying not in underlying_summary:
                underlying_summary[underlying] = {
                    'positions': [],
                    'total_contracts': 0,
                    'total_lots': 0,
                    'lot_size': position.lot_size,  # Store lot size
                    'system_price': position.underlying_price
                }
            underlying_summary[underlying]['positions'].append(position)
            underlying_summary[underlying]['total_contracts'] += 1
            underlying_summary[underlying]['total_lots'] += abs(position.position)
        
        # Write summary data
        current_row = 2
        for underlying in sorted(underlying_summary.keys()):
            data = underlying_summary[underlying]
            
            # Underlying name
            ws.cell(row=current_row, column=1, value=underlying)
            
            # Contract count and lots
            ws.cell(row=current_row, column=2, value=data['total_contracts'])
            ws.cell(row=current_row, column=3, value=data['total_lots'])
            ws.cell(row=current_row, column=4, value=data['lot_size'])
            
            # System deliverable (sum from all positions)
            system_deliverable = sum(pos.deliverable for pos in data['positions'])
            ws.cell(row=current_row, column=5, value=system_deliverable)
            
            # Formulas for Override and BBG deliverables (referencing Master sheet)
            # These will pull from the grouped sheets
            ws.cell(row=current_row, column=6, value=f"=SUMIF(Master_All_Expiries!A:A,A{current_row},Master_All_Expiries!H:H)")
            ws.cell(row=current_row, column=7, value=f"=SUMIF(Master_All_Expiries!A:A,A{current_row},Master_All_Expiries!L:L)")
            
            # Prices
            ws.cell(row=current_row, column=8, value=data['system_price'])
            ws.cell(row=current_row, column=9, value="")  # Override price (manual input)
            ws.cell(row=current_row, column=10, value=f'=@BDP(A{current_row},"PX_LAST")')  # BBG price
            
            # Highlight row based on deliverable size
            if abs(system_deliverable) > 100:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )
            
            current_row += 1
        
        # Add totals row
        total_row = current_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
        
        # Total formulas
        ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{current_row-1})")
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{current_row-1})")
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{current_row-1})")
        ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{current_row-1})")
        ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{current_row-1})")
        
        # Style total row
        for col in range(1, 11):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.font = Font(bold=True)
        
        # Auto-size columns
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Freeze top row
        ws.freeze_panes = ws['A2']
    
    def _create_price_alert_sheet(self, workbook, positions: List) -> None:
        """Create price alert sheet for positions near strike price"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
        
        ws = workbook.create_sheet(title="Price_Alerts", index=1)
        
        # Add threshold input cell
        ws.cell(row=1, column=1, value="Alert Threshold (%):")
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2, value=1.0)  # Default 1%
        ws.cell(row=1, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Headers
        headers = [
            'Underlying', 'Symbol', 'Type', 'Strike', 'Current Price',
            'Moneyness %', 'Days to Expiry', 'Position', 'Alert Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        
        # Filter for options only
        options_positions = [pos for pos in positions if pos.security_type in ['Call', 'Put']]
        
        current_row = 4
        today = datetime.now()
        
        for position in sorted(options_positions, key=lambda x: (x.underlying_ticker, x.expiry)):
            # Basic info
            ws.cell(row=current_row, column=1, value=position.underlying_ticker)
            ws.cell(row=current_row, column=2, value=position.bloomberg_ticker)
            ws.cell(row=current_row, column=3, value=position.security_type)
            ws.cell(row=current_row, column=4, value=position.strike)
            
            # Current price (formula to reference Net Position Summary)
            price_formula = f"=VLOOKUP(A{current_row},Net_Position_Summary!A:H,8,FALSE)"
            ws.cell(row=current_row, column=5, value=price_formula)
            
            # Moneyness % calculation
            if position.security_type == "Call":
                moneyness_formula = f"=IF(E{current_row}>0,(E{current_row}-D{current_row})/D{current_row}*100,0)"
            else:  # Put
                moneyness_formula = f"=IF(E{current_row}>0,(D{current_row}-E{current_row})/D{current_row}*100,0)"
            ws.cell(row=current_row, column=6, value=moneyness_formula)
            
            # Days to expiry
            days_to_expiry = (position.expiry - today).days
            ws.cell(row=current_row, column=7, value=days_to_expiry)
            
            # Position
            ws.cell(row=current_row, column=8, value=position.position)
            
            # Alert Status formula
            alert_formula = f'=IF(ABS(F{current_row})<=$B$1,"🔴 NEAR STRIKE",IF(ABS(F{current_row})<=($B$1*2),"🟡 WATCH","🟢 SAFE"))'
            ws.cell(row=current_row, column=9, value=alert_formula)
            
            # Conditional formatting for alert rows
            if position.underlying_price:
                moneyness = abs((position.underlying_price - position.strike) / position.strike * 100)
                if moneyness <= 1.0:
                    for col in range(1, 10):
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"
                        )
                elif moneyness <= 2.0:
                    for col in range(1, 10):
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                        )
            
            # Highlight if expiring soon (within 7 days)
            if days_to_expiry <= 7:
                ws.cell(row=current_row, column=7).fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                ws.cell(row=current_row, column=7).font = Font(color="FFFFFF", bold=True)
            
            current_row += 1
        
        # Add summary statistics
        summary_row = current_row + 2
        ws.cell(row=summary_row, column=1, value="SUMMARY")
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Total Options:")
        ws.cell(row=summary_row, column=2, value=len(options_positions))
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Positions < 1% from strike:")
        ws.cell(row=summary_row, column=2, value=f'=COUNTIF(I4:I{current_row-1},"*NEAR*")')
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Expiring within 7 days:")
        ws.cell(row=summary_row, column=2, value=f'=COUNTIF(G4:G{current_row-1},"<=7")')
        
        # Auto-size columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze header rows
        ws.freeze_panes = ws['A4']
    
    def _create_unmapped_sheet(self, workbook) -> None:
        """Create sheet with unmapped symbols for mapping updates"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title="Unmapped_Symbols")
        
        # Headers for unmapped positions
        headers = [
            'Symbol', 'Contract ID', 'Position', 'Lot Size', 
            'Series', 'Expiry', 'Strike', 'Option Type', 
            'Row Number', 'Suggested Ticker', 'Suggested Cash'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Add unmapped positions
        unique_symbols = {}
        current_row = 2
        
        for unmapped in self.unmapped_positions:
            symbol = unmapped['symbol']
            
            # Track unique symbols
            if symbol not in unique_symbols:
                unique_symbols[symbol] = []
            unique_symbols[symbol].append(unmapped)
            
            # Write position details
            ws.cell(row=current_row, column=1, value=symbol)
            ws.cell(row=current_row, column=2, value=unmapped.get('contract_id', ''))
            ws.cell(row=current_row, column=3, value=unmapped.get('position', 0))
            ws.cell(row=current_row, column=4, value=unmapped.get('lot_size', 1))
            ws.cell(row=current_row, column=5, value=unmapped.get('series', ''))
            ws.cell(row=current_row, column=6, value=unmapped.get('expiry', '').strftime('%Y-%m-%d') if hasattr(unmapped.get('expiry'), 'strftime') else str(unmapped.get('expiry', '')))
            ws.cell(row=current_row, column=7, value=unmapped.get('strike', 0))
            ws.cell(row=current_row, column=8, value=unmapped.get('option_type', ''))
            ws.cell(row=current_row, column=9, value=unmapped.get('row_number', ''))
            
            # Suggested mapping (to be filled by user)
            ws.cell(row=current_row, column=10, value="")  # Suggested ticker
            ws.cell(row=current_row, column=11, value="")  # Suggested cash ticker
            
            # Highlight row
            for col in range(1, 12):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            current_row += 1
        
        # Add summary section
        current_row += 2
        ws.cell(row=current_row, column=1, value="SUMMARY")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Unique Unmapped Symbols:")
        ws.cell(row=current_row, column=2, value=len(unique_symbols))
        
        current_row += 1
        ws.cell(row=current_row, column=1, value="Total Unmapped Positions:")
        ws.cell(row=current_row, column=2, value=len(self.unmapped_positions))
        
        current_row += 2
        ws.cell(row=current_row, column=1, value="UNIQUE SYMBOLS LIST")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        current_row += 1
        for symbol in sorted(unique_symbols.keys()):
            ws.cell(row=current_row, column=1, value=symbol)
            ws.cell(row=current_row, column=2, value=f"{len(unique_symbols[symbol])} positions")
            current_row += 1
        
        # Add instructions
        current_row += 2
        instructions = [
            "INSTRUCTIONS:",
            "1. Copy the unique symbols from this sheet",
            "2. Add them to your 'futures mapping.csv' file with correct ticker mappings",
            "3. Re-run the transformation to include these positions",
            "",
            "MAPPING FILE FORMAT:",
            "Symbol,Ticker,Cash",
            "LTIM,LTIM,LTIM IS Equity",
            "RELIANCE,RELIANCE,RELIANCE IS Equity"
        ]
        
        for instruction in instructions:
            ws.cell(row=current_row, column=1, value=instruction)
            if instruction.startswith("INSTRUCTIONS:") or instruction.startswith("MAPPING FILE FORMAT:"):
                ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        
        # Auto-size columns
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_grouped_sheet(self, workbook, sheet_name: str, positions: List) -> None:
        """Create sheet with Excel row grouping and formulas"""
        
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title=sheet_name)
        
        # Headers with BBG Price and BBG Deliverable
        headers = [
            'Underlying', 'Symbol', 'Expiry', 'Position', 'Type', 'Strike', 
            'System Deliverable', 'Override Deliverable', 'System Price', 'Override Price', 
            'BBG Price', 'BBG Deliverable'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Group positions by underlying ticker
        grouped_positions = {}
        for position in positions:
            underlying = position.underlying_ticker
            if underlying not in grouped_positions:
                grouped_positions[underlying] = []
            grouped_positions[underlying].append(position)
        
        current_row = 2
        group_ranges = []
        
        for underlying_ticker in sorted(grouped_positions.keys()):
            positions_group = grouped_positions[underlying_ticker]
            underlying_row = current_row
            
            # Write underlying summary row
            ws.cell(row=underlying_row, column=1, value=underlying_ticker)
            
            # Get representative position for prices
            repr_position = positions_group[0]
            system_price = repr_position.underlying_price
            if system_price is not None:
                system_price = round(system_price, 1)
            ws.cell(row=underlying_row, column=9, value=system_price)  # System Price (Column I)
            
            # BBG Price formula
            bbg_formula = f'=@BDP(A{underlying_row},"PX_LAST")'
            ws.cell(row=underlying_row, column=11, value=bbg_formula)  # BBG Price (Column K)
            
            # Style underlying summary row
            for col in range(1, 13):  # 12 columns
                cell = ws.cell(row=underlying_row, column=col)
                cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                cell.font = Font(bold=True, size=11)
            
            current_row += 1
            detail_start_row = current_row
            
            # Write individual positions
            for position in sorted(positions_group, key=lambda x: x.expiry):
                # Basic data
                ws.cell(row=current_row, column=2, value=position.bloomberg_ticker)
                ws.cell(row=current_row, column=3, value=position.expiry.strftime('%Y-%m-%d'))
                ws.cell(row=current_row, column=4, value=position.position)
                ws.cell(row=current_row, column=5, value=position.security_type)
                ws.cell(row=current_row, column=6, value=position.strike if position.strike > 0 else None)
                
                # Link prices to underlying row
                ws.cell(row=current_row, column=9, value=f"=I{underlying_row}")   # System Price
                ws.cell(row=current_row, column=10, value=f"=J{underlying_row}")  # Override Price
                ws.cell(row=current_row, column=11, value=f"=K{underlying_row}")  # BBG Price
                
                # DELIVERABLE FORMULAS
                system_formula = self._make_system_formula(current_row, underlying_row, position)
                override_formula = self._make_override_formula(current_row, underlying_row, position)
                bbg_formula = self._make_bbg_formula(current_row, underlying_row, position)
                
                ws.cell(row=current_row, column=7, value=system_formula)    # System Deliverable
                ws.cell(row=current_row, column=8, value=override_formula)  # Override Deliverable
                ws.cell(row=current_row, column=12, value=bbg_formula)      # BBG Deliverable
                
                current_row += 1
            
            detail_end_row = current_row - 1
            
            # Add total formulas to underlying row
            ws.cell(row=underlying_row, column=7, value=f"=SUM(G{detail_start_row}:G{detail_end_row})")   # System Total
            ws.cell(row=underlying_row, column=8, value=f"=SUM(H{detail_start_row}:H{detail_end_row})")   # Override Total
            ws.cell(row=underlying_row, column=12, value=f"=SUM(L{detail_start_row}:L{detail_end_row})")  # BBG Total
            
            # Store group range for later processing
            if detail_end_row >= detail_start_row:
                group_ranges.append((detail_start_row, detail_end_row))
        
        # Apply grouping - MUST be done after all data is written
        if group_ranges:
            try:
                # Set outline properties BEFORE creating groups
                ws.sheet_properties.outlinePr.summaryBelow = False
                ws.sheet_properties.outlinePr.summaryRight = False
                
                # Apply grouping to each range
                for start_row, end_row in group_ranges:
                    if end_row > start_row:
                        # Multiple rows - create a group
                        ws.row_dimensions.group(start_row, end_row, hidden=True, outline_level=1)
                    elif end_row == start_row:
                        # Single row - just hide it without grouping
                        ws.row_dimensions[start_row].hidden = True
                
                logger.info(f"✅ Created {len(group_ranges)} row groups in sheet '{sheet_name}'")
                
            except Exception as e:
                logger.warning(f"Could not create row groups: {str(e)}")
        
        # Auto-size columns
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _make_system_formula(self, current_row: int, underlying_row: int, position) -> str:
        """Create system deliverable formula using Column I (system price)"""
        if position.security_type == "Futures":
            return f"=D{current_row}"
        elif position.security_type == "Call":
            return f'=IF(AND(I{underlying_row}>0,F{current_row}>0),IF(I{underlying_row}>F{current_row},D{current_row},0),D{current_row})'
        elif position.security_type == "Put":
            return f'=IF(AND(I{underlying_row}>0,F{current_row}>0),IF(I{underlying_row}<F{current_row},-D{current_row},0),-D{current_row})'
        return "=0"
    
    def _make_override_formula(self, current_row: int, underlying_row: int, position) -> str:
        """Create override deliverable formula using Column J (manual input)"""
        if position.security_type == "Futures":
            return f"=D{current_row}"
        elif position.security_type == "Call":
            return f'=IF(J{underlying_row}>0,IF(J{underlying_row}>F{current_row},D{current_row},0),G{current_row})'
        elif position.security_type == "Put":
            return f'=IF(J{underlying_row}>0,IF(J{underlying_row}<F{current_row},-D{current_row},0),G{current_row})'
        return f"=G{current_row}"
    
    def _make_bbg_formula(self, current_row: int, underlying_row: int, position) -> str:
        """Create BBG deliverable formula using Column K (BBG price input)"""
        if position.security_type == "Futures":
            return f"=D{current_row}"
        elif position.security_type == "Call":
            return f'=IF(K{underlying_row}>0,IF(K{underlying_row}>F{current_row},D{current_row},0),G{current_row})'
        elif position.security_type == "Put":
            return f'=IF(K{underlying_row}>0,IF(K{underlying_row}<F{current_row},-D{current_row},0),G{current_row})'
        return f"=G{current_row}"
    
    def get_summary_stats(self) -> Dict:
        """Get summary statistics including unmapped positions"""
        if not self.positions and not self.unmapped_positions:
            return {
                'total_positions': 0,
                'total_underlyings': 0,
                'total_deliverables': 0,
                'positions_by_type': {},
                'underlyings_list': [],
                'input_format': self.input_format or 'unknown',
                'unmapped_count': 0,
                'unmapped_symbols': []
            }
        
        total_positions = len(self.positions)
        total_deliverables = sum(abs(pos.deliverable) for pos in self.positions)
        underlyings = set(pos.underlying_ticker for pos in self.positions)
        
        by_type = {}
        for pos in self.positions:
            if pos.security_type not in by_type:
                by_type[pos.security_type] = 0
            by_type[pos.security_type] += 1
        
        # Get unique unmapped symbols
        unmapped_symbols = list(set(pos['symbol'] for pos in self.unmapped_positions))
        
        return {
            'total_positions': total_positions,
            'total_underlyings': len(underlyings),
            'total_deliverables': total_deliverables,
            'positions_by_type': by_type,
            'underlyings_list': sorted(underlyings),
            'input_format': self.input_format,
            'unmapped_count': len(self.unmapped_positions),
            'unmapped_symbols': sorted(unmapped_symbols)
        }
    
    # Helper methods
    def _parse_date(self, date_value) -> datetime:
        """Parse various date formats"""
        if pd.isna(date_value):
            return datetime.now()
        
        if isinstance(date_value, datetime):
            return date_value
        
        if isinstance(date_value, (int, float)):
            # Excel serial date
            try:
                return pd.to_datetime('1900-01-01') + pd.Timedelta(days=date_value-2)
            except:
                return datetime.now()
        
        if isinstance(date_value, str):
            try:
                return pd.to_datetime(date_value)
            except:
                return datetime.now()
        
        return datetime.now()
    
    def _generate_bloomberg_ticker(self, symbol: str, fo_ticker: str, series: str, 
                                 expiry: datetime, strike: float, option_type: str) -> str:
        """Generate Bloomberg ticker format"""
        try:
            is_future = series == 'FUTSTK' or option_type == 'FF'
            
            if is_future:
                # Futures format
                month_code = self._get_month_code(expiry.month)
                year_digit = str(expiry.year)[-1]
                return f"{fo_ticker}={month_code}{year_digit} IS Equity"
            else:
                # Options format
                expiry_str = expiry.strftime('%m/%d/%y')
                
                if option_type in ['CE', 'C']:
                    option_char = 'C'
                elif option_type in ['PE', 'P']:
                    option_char = 'P'
                else:
                    option_char = 'C'  # Default
                
                strike_str = str(int(strike)) if strike > 0 else '0'
                
                return f"{fo_ticker} IS {expiry_str} {option_char}{strike_str} Equity"
                
        except Exception as e:
            logger.warning(f"Error generating Bloomberg ticker for {symbol}: {str(e)}")
            return f"{fo_ticker} IS Equity"
    
    def _get_month_code(self, month: int) -> str:
        """Get futures month code"""
        month_codes = {
            1: 'F', 2: 'G', 3: 'H', 4: 'J', 5: 'K', 6: 'M',
            7: 'N', 8: 'Q', 9: 'U', 10: 'V', 11: 'X', 12: 'Z'
        }
        return month_codes.get(month, 'Z')
    
    def _get_security_type(self, series: str, option_type: str) -> str:
        """Determine security type"""
        if series == 'FUTSTK' or option_type == 'FF':
            return 'Futures'
        elif option_type in ['CE', 'C']:
            return 'Call'
        elif option_type in ['PE', 'P']:
            return 'Put'
        else:
            return 'Unknown'
    
    def _is_in_the_money(self, option_type: str, strike: float, underlying_price: float) -> bool:
        """Determine if option is in-the-money"""
        if option_type in ['CE', 'Call', 'C']:
            return underlying_price > strike
        elif option_type in ['PE', 'Put', 'P']:
            return underlying_price < strike
        else:
            return False


def send_email_report(output_file: str, stats: Dict, recipients: List[str]) -> bool:
    """Send email report with Excel attachment"""
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        import os
        
        # Email configuration (you'll need to set these)
        SMTP_SERVER = "smtp.gmail.com"  # Change for your email provider
        SMTP_PORT = 587
        SENDER_EMAIL = os.environ.get('SENDER_EMAIL', '')  # Set environment variable
        SENDER_PASSWORD = os.environ.get('SENDER_PASSWORD', '')  # Set environment variable
        
        if not SENDER_EMAIL or not SENDER_PASSWORD:
            logger.warning("Email credentials not configured. Set SENDER_EMAIL and SENDER_PASSWORD environment variables.")
            return False
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = f"Portfolio Transformation Report - {datetime.now().strftime('%Y-%m-%d')}"
        
        # Email body
        body = f"""
Portfolio Transformation Report

Summary:
- Input Format: {stats['input_format']}
- Total Positions Processed: {stats['total_positions']}
- Total Underlyings: {stats['total_underlyings']}
- Positions by Type: {stats['positions_by_type']}

Unmapped Symbols: {stats.get('unmapped_count', 0)}
{f"Symbols needing mapping: {', '.join(stats.get('unmapped_symbols', [])[:5])}" if stats.get('unmapped_count', 0) > 0 else ''}

The detailed Excel report is attached.

Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach Excel file
        with open(output_file, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(output_file)}'
            )
            msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        logger.info(f"✅ Email sent successfully to {', '.join(recipients)}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        return False


def main():
    """Main function with streamlined workflow"""
    try:
        print("🚀 UNIVERSAL PORTFOLIO TRANSFORMER - ENHANCED VERSION")
        print("=" * 60)
        print("✅ Supports 3 input formats with smart detection")
        print("✅ Flexible start row detection for all formats")
        print("✅ Auto-loads mapping")
        print("✅ Auto-fetches prices")
        print("✅ Bloomberg integration")
        print("✅ Password protection")
        print("✅ Net position summary")
        print("✅ Price alerts for options")
        print("=" * 60)
        
        # Step 1: Select Fund
        fund_name = select_fund()
        
        # Step 2: Select input file
        input_file = select_file_from_directory("position data", ['.csv', '.xlsx', '.xls'])
        
        # Step 3: Generate output filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"{fund_name}_{timestamp}.xlsx"
        
        # Step 4: Initialize transformer
        print(f"\n📄 PROCESSING {fund_name} PORTFOLIO...")
        transformer = EnhancedPortfolioTransformer(fund_name)
        
        # Step 5: Load mapping (auto)
        print("📋 Loading futures mapping...")
        transformer.load_mapping_data()
        
        # Step 6: Load positions (auto-detect format with improved logic)
        print("📋 Loading positions (smart format detection)...")
        transformer.load_positions(input_file)
        
        # Step 7: Calculate deliverables (auto-fetch prices)
        print("🧮 Calculating deliverables and fetching prices...")
        transformer.calculate_deliverables(auto_fetch_prices=True)
        
        # Step 8: Save output
        print("💾 Generating Excel output...")
        transformer.save_output_excel(output_file)
        
        # Step 9: Print summary
        stats = transformer.get_summary_stats()
        print(f"\n🎉 TRANSFORMATION COMPLETED SUCCESSFULLY!")
        print(f"   📁 Fund: {fund_name}")
        print(f"   📊 Input format: {stats['input_format']}")
        print(f"   📈 Total positions: {stats['total_positions']}")
        print(f"   🏢 Total underlyings: {stats['total_underlyings']}")
        print(f"   📋 Positions by type: {stats['positions_by_type']}")
        
        # Show unmapped symbols if any
        if stats.get('unmapped_count', 0) > 0:
            print(f"\n   ⚠️ UNMAPPED SYMBOLS: {stats['unmapped_count']} positions")
            print(f"   🔍 Unique symbols without mapping: {len(stats['unmapped_symbols'])}")
            print(f"   📝 Check 'Unmapped_Symbols' sheet in output file")
            print(f"   Unmapped symbols: {', '.join(stats['unmapped_symbols'][:10])}")
            if len(stats['unmapped_symbols']) > 10:
                print(f"   ... and {len(stats['unmapped_symbols']) - 10} more")
        
        print(f"   💾 Output saved to: {output_file}")
        
        # Step 10: Ask about email
        send_email = input("\n📧 Send email report? (y/n): ").strip().lower()
        if send_email == 'y':
            recipients_input = input("Enter email addresses (comma-separated): ").strip()
            if recipients_input:
                recipients = [email.strip() for email in recipients_input.split(',')]
                print("📧 Sending email report...")
                if send_email_report(output_file, stats, recipients):
                    print("✅ Email sent successfully!")
                else:
                    print("❌ Email sending failed. Check configuration.")
        
        print(f"\n🔧 FEATURES INCLUDED:")
        print(f"   ✅ Smart format detection: {stats['input_format']}")
        print(f"   ✅ Yahoo Finance prices fetched")
        print(f"   ✅ Bloomberg formulas: =@BDP(underlying,\"PX_LAST\")")
        print(f"   ✅ Excel formulas for deliverable calculations")
        print(f"   ✅ Collapsible grouped rows by underlying")
        print(f"   ✅ Net position summary sheet")
        print(f"   ✅ Price alert sheet with configurable threshold")
        
        # Show sample positions
        if transformer.positions:
            print(f"\n📊 SAMPLE PROCESSED POSITIONS:")
            for i, pos in enumerate(transformer.positions[:3]):
                print(f"   {i+1}. {pos.symbol} ({pos.security_type}) - {pos.position} lots (Lot Size: {pos.lot_size})")
        
    except KeyboardInterrupt:
        print("\n❌ Transformation cancelled by user")
    except SystemExit:
        print("❌ Transformation stopped")
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        logger.error(f"Error in main execution: {str(e)}")
        import traceback
        traceback.print_exc()
        
    except KeyboardInterrupt:
        print("\n❌ Transformation cancelled by user")
    except SystemExit:
        print("❌ Transformation stopped")
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        logger.error(f"Error in main execution: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
